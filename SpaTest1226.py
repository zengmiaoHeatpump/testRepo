#!/usr/bin/python
# -*- coding:utf-8 -*-
import time
import minimalmodbus
from openpyxl import Workbook
from openpyxl import load_workbook
import pymysql
import sys
import os
import netifaces as ni
picdir = os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), 'pic')
libdir = os.path.join(os.path.dirname(os.path.dirname(os.path.realpath(__file__))), 'lib')
if os.path.exists(libdir):
    sys.path.append(libdir)

import logging
from waveshare_epd import epd7in5bc
import time
from PIL import Image,ImageDraw,ImageFont
import traceback
# sensor modbus address
PowerMeasureAddress = 31
IndoorTemAddress = 1
OutDoorTemAddress = 9
IDU_1_LowAddress = 21
IDU_1_MidAddress = 22
IDU_1_HighAddress = 23
IDU_2_LowAddress = 25
IDU_2_MidAddress = 24
IDU_2_HighAddress = 27
instList = []
power_TRH_List = []
FN = "/home/pi/Spatest/TRH"+time.strftime("%Y%m%d%H%M", time.localtime())+".xlsx"
SHEET_TITLE="Power_T_RH"
SHEET_1ST=("No","Datetime","Power input","Power Summary","Outdoor Tem","Outdoor RH","IDU_1_Low_Tem",
    "IDU_1_Mid_Tem","IDU_1_Mid_RH","IDU_1_High_Tem","IDU_2_Low_Tem","IDU_2_Mid_Tem","IDU_2_Mid_RH",
    "IDU_2_High_Tem","Room Tem","Room RH")


###################set record time ###################
MAXtime = 2400000
STEP = 300
######################################################
def showip() :        
    ni.ifaddresses('eth0')
    lan = ni.ifaddresses('eth0')
    if ni.AF_INET in lan:
        lanip = ni.ifaddresses('eth0')[ni.AF_INET][0]['addr']
        return ('eth0:'+lanip )       
    ni.ifaddresses('wlan0')
    wlan = ni.ifaddresses('wlan0')
    if ni.AF_INET in wlan:
        wlanip = ni.ifaddresses('wlan0')[ni.AF_INET][0]['addr']
        return ('wlan0:'+wlanip)
    return 'disconnect'  

def initSensor():
    # init power measure
    inst_PowerMeasure = minimalmodbus.Instrument(
        '/dev/ttyAMA0', PowerMeasureAddress)
    instList.append(inst_PowerMeasure)

    # init outdoor tem
    inst_OutDoorTem = minimalmodbus.Instrument(
        '/dev/ttyAMA0', OutDoorTemAddress)
    instList.append(inst_OutDoorTem)
    # init IDU_1_Low
    inst_IDU_1_Low = minimalmodbus.Instrument('/dev/ttyAMA0', IDU_1_LowAddress)
    instList.append(inst_IDU_1_Low)
    # init IDU_1_Mid
    inst_IDU_1_Mid = minimalmodbus.Instrument('/dev/ttyAMA0', IDU_1_MidAddress)
    instList.append(inst_IDU_1_Mid)
    # init IDU_1_High
    inst_IDU_1_High = minimalmodbus.Instrument(
        '/dev/ttyAMA0', IDU_1_HighAddress)
    instList.append(inst_IDU_1_High)
    # init IDU_2_Low
    inst_IDU_2_Low = minimalmodbus.Instrument('/dev/ttyAMA0', IDU_2_LowAddress)
    instList.append(inst_IDU_2_Low)
    # init IDU_2_Mid
    inst_IDU_2_Mid = minimalmodbus.Instrument('/dev/ttyAMA0', IDU_2_MidAddress)
    instList.append(inst_IDU_2_Mid)
    # init IDU_2_High
    inst_IDU_2_High = minimalmodbus.Instrument(
        '/dev/ttyAMA0', IDU_2_HighAddress)
    instList.append(inst_IDU_2_High)
    # init indoor tem
    inst_IndoorTem = minimalmodbus.Instrument('/dev/ttyAMA0', IndoorTemAddress)
    instList.append(inst_IndoorTem)

    for inst in instList:
        inst.serial.baudrate = 9600
        inst.serial.timeout = 1
        inst.serial.stopbits = 1

    instList[0].serial.stopbits = 2  # powermeasure stopbits =2
    ############init xls file#############
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = SHEET_TITLE
    sheet1.append(SHEET_1ST)
    wb.save(FN)
    
    

##################仅仅读取一行#############################
def read_All_Data():
    power_TRH_List=[]    
    recordTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    power_TRH_List.append(recordTime)  # No2
    ######################
    MQHDpower =round(instList[0].read_float(8196, 3), 2)
    MQHDpowerSum =round(instList[0].read_float(16384, 3), 2)
    time.sleep(1)
    power_TRH_List.append(MQHDpower)  # No3
    power_TRH_List.append(MQHDpowerSum)  # No4
    #########################
    odu_tem = instList[1].read_register(0, 1, 3, signed=True)
    odu_rh = instList[1].read_register(1, 1, 3)
    time.sleep(1)
    power_TRH_List.append(odu_tem)  # No5
    power_TRH_List.append(odu_rh)  # No6
    ####################
    idu_1_low_tem = instList[2].read_register(0, 1, 3, signed=True)
    time.sleep(1)
    power_TRH_List.append(idu_1_low_tem)  # No7
    idu_1_mid_tem = instList[3].read_register(0, 1, 3, signed=True)
    idu_1_mid_rh = instList[3].read_register(1, 1, 3)
    power_TRH_List.append(idu_1_mid_tem)  # No8
    power_TRH_List.append(idu_1_mid_rh)  # No9
    time.sleep(1)
    idu_1_high_tem = instList[4].read_register(0, 1, 3, signed=True)
    time.sleep(1)
    power_TRH_List.append(idu_1_high_tem)  # No10
    ##########################
    idu_2_low_tem = instList[5].read_register(0, 1, 3, signed=True)
    time.sleep(1)
    power_TRH_List.append(idu_2_low_tem)  # No11
    idu_2_mid_tem = instList[6].read_register(0, 1, 3, signed=True)
    idu_2_mid_rh = instList[6].read_register(1, 1, 3)
    power_TRH_List.append(idu_2_mid_tem)  # No12
    power_TRH_List.append(idu_2_mid_rh)  # No13
    time.sleep(1)
    idu_2_high_tem = instList[7].read_register(0, 1, 3, signed=True)
    time.sleep(1)
    power_TRH_List.append(idu_2_high_tem)  # No14
    ########################
    room_tem = instList[8].read_register(0, 1, 3, signed=True)
    room_rh = instList[8].read_register(1, 1, 3)
    time.sleep(1)
    power_TRH_List.append(room_tem)  # No15
    power_TRH_List.append(room_rh)  # No16
    # inst.read_register(0,1,3,signed=True)
    return power_TRH_List

def EinkDisplay(dataList):
    logging.basicConfig(level=logging.DEBUG)            
    logging.info("AAT test ")    
    epd = epd7in5bc.EPD()
    logging.info("init and Clear")
    epd.init()
    epd.Clear()
    time.sleep(1)    
    # Drawing on the image
    logging.info("Drawing")    
    #font24 = ImageFont.truetype(os.path.join(picdir, 'Font.ttc'), 24)
    font18 = ImageFont.truetype(os.path.join(picdir, 'Font.ttc'), 18)
    font12 = ImageFont.truetype(os.path.join(picdir, 'Font.ttc'), 12)
    font9 = ImageFont.truetype(os.path.join(picdir, 'Font.ttc'), 9)
    time.sleep(1)        
    logging.info("1.read bmp file")
    HBlackimage = Image.open(os.path.join(picdir, 'aatblack.bmp'))
    HRYimage = Image.new('1', (epd.width, epd.height), 255)
    drawblack = ImageDraw.Draw(HBlackimage)
    drawblack.text((30, 30), 'AAT conference room test system', font = font18, fill = 0)
    drawblack.text((30, 100), 'Datetime'+str(dataList[0]), font = font12, fill = 0)
    drawblack.text((30, 150), 'OutDoor          ODU', font = font12, fill = 0)
    drawblack.text((105, 165), 'power comsumption', font = font9, fill = 0)
    drawblack.text((30, 185), 'T:'+str(dataList[3])+u'C'+'               Realtime:'+str(dataList[1])+'Kw', font = font12, fill = 0)
    drawblack.text((30, 225), 'RH:'+str(dataList[4])+'%'+'         Sum:'+str(dataList[2])+'KWh', font = font12, fill = 0)
    drawblack.text((275, 170), 'No.2                                 No.1                Indoor', font = font12, fill = 0)
    drawblack.text((275, 215), 'Th:'+str(dataList[12])+u'C'+'                        Th:'+str(dataList[8])+u'C'+'           T:'+str(dataList[13])+u'C', font = font12, fill = 0)
    drawblack.text((275, 240), 'Tm:'+str(dataList[10])+u'C'+'                      Tm:'+str(dataList[6])+u'C'+'          RH:'+str(dataList[14])+'%', font = font12, fill = 0)
    drawblack.text((275, 260), 'TL:'+str(dataList[9])+u'C'+'                        TL:'+str(dataList[5])+'C', font = font12, fill = 0)
    drawblack.text((540, 80), 'Scan QR Code', font = font12, fill = 0)
    drawblack.text((540, 95), 'for historical data', font = font12, fill = 0)
    #HBlackimage.rotate(180)
    epd.display(epd.getbuffer(HBlackimage), epd.getbuffer(HRYimage))
    logging.info("Goto Sleep...")
    #epd.sleep()        
    #time.sleep(60)

def saveToXls(datalist,counter):
    #wb = Workbook()
    wb = load_workbook(filename = FN)
    sheet1 = wb.active
    #sheet1.title = SHEET_TITLE      
    xlsList=datalist 
    xlsList.insert(0,counter)
    sheet1.append(xlsList)
    wb.save(FN)   

def uploadToDatabase(datalist):
    Tecent_DATABASE = pymysql.connect("49.235.6.254","zengmiao","airwell123","trhData" )
    cursor = Tecent_DATABASE.cursor()
    tupleData=tuple(datalist)        
    sql00 = "INSERT INTO dataShow_trh(record_date,ODU_Powerinput,ODU_PowerSum,ODU_tem,ODU_rh,IDU_1_Low_tem,IDU_1_mid_tem,IDU_1_mid_rh,IDU_1_high_tem,IDU_2_Low_tem,IDU_2_mid_tem,IDU_2_mid_rh,IDU_2_high_tem,Room_tem,Room_rh) VALUES "
    #val="'%s','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f','%.2f'" %tupleData
    val="('{}', '{:.2f}','{:.2f}', '{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}','{:.2f}')".format(*tupleData)
    sql=sql00+val
    print(sql)
    
    try:
    # 执行sql语句
        Tecent_DATABASE.ping(reconnect=True)
        cursor.execute(sql)
    # 提交到数据库执行
        Tecent_DATABASE.commit()
    except:
    # 如果发生错误则回滚
        Tecent_DATABASE.rollback() 
        # 关闭数据库连接
        Tecent_DATABASE.close()

def update_xls_SQL():

    ######################################
    counter = 1
    while counter < MAXtime:
        #################
        power_TRH_List=read_All_Data()
        #print(power_TRH_List)
        ################################
        uploadToDatabase(power_TRH_List)        
        EinkDisplay(power_TRH_List)
        saveToXls(power_TRH_List,counter)
        counter=counter+1
        time.sleep(STEP)
        ##############

#########################################
initSensor()
update_xls_SQL()
